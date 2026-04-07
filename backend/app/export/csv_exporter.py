"""Export leads to CSV and Excel formats."""

import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.business import Business


EXPORT_COLUMNS = [
    ("Nombre", "name"),
    ("Teléfono", "phone"),
    ("Teléfono E.164", "phone_e164"),
    ("Email", "email"),
    ("Dirección", "address"),
    ("Ciudad", "city"),
    ("Provincia", "province"),
    ("CP", "postal_code"),
    ("País", "country_code"),
    ("Web", "website_url"),
    ("Estado Web", "website_status"),
    ("SSL", "website_ssl"),
    ("Mobile", "website_mobile"),
    ("CMS", "website_cms"),
    ("Rating Google", "google_rating"),
    ("Reseñas", "google_reviews"),
    ("Google Maps", "google_maps_url"),
    ("Categorías", "categories"),
    ("Score", "lead_score"),
    ("Robinson", "lista_robinson"),
    ("Fuente", "data_source"),
]


def export_csv(businesses: list[Business]) -> str:
    """Export businesses to CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([col[0] for col in EXPORT_COLUMNS])

    # Data rows
    for biz in businesses:
        row = []
        for _, attr in EXPORT_COLUMNS:
            val = getattr(biz, attr, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            elif val is None:
                val = ""
            row.append(str(val))
        writer.writerow(row)

    return output.getvalue()


def export_excel(businesses: list[Business], sheet_name: str = "Leads") -> bytes:
    """Export businesses to Excel (.xlsx) with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    score_high_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    score_med_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    score_low_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    # Header row
    for col_idx, (col_name, _) in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, biz in enumerate(businesses, 2):
        for col_idx, (_, attr) in enumerate(EXPORT_COLUMNS, 1):
            val = getattr(biz, attr, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            elif val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            # Conditional formatting for score column
            if attr == "lead_score" and isinstance(val, int):
                if val >= 60:
                    cell.fill = score_high_fill
                elif val >= 30:
                    cell.fill = score_med_fill
                else:
                    cell.fill = score_low_fill

    # Auto-width columns
    for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(ws.max_row + 1, 100))
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
